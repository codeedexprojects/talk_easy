from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.views import APIView
from .models import *
from .serializers import *
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from executives.authentication import ExecutiveTokenAuthentication
from django.utils.timezone import now
from rest_framework.permissions import IsAdminUser
import razorpay
from django.conf import settings
from accounts.pagination import *
from executives.permissions import IsAdminUser
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal


#  Category Create & List
class RechargePlanCategoryListCreateAPIView(generics.ListCreateAPIView):
    queryset = RechargePlanCatogary.objects.filter(is_deleted=False)
    serializer_class = RechargePlanCategorySerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

#  Category Detail
class RechargePlanCategoryDetailAPIView(generics.RetrieveUpdateAPIView):
    queryset = RechargePlanCatogary.objects.filter(is_deleted=False)
    serializer_class = RechargePlanCategorySerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]


#  Category Soft Delete
class RechargePlanCategoryDeleteAPIView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def delete(self, request, pk):
        try:
            category = RechargePlanCatogary.objects.get(pk=pk, is_deleted=False)
            category.is_deleted = True
            category.save()
            return Response({"message": "Category deleted successfully"}, status=status.HTTP_200_OK)
        except RechargePlanCatogary.DoesNotExist:
            return Response({"error": "Category not found"}, status=status.HTTP_404_NOT_FOUND)


#  Plan Create & List
class RechargePlanListCreateAPIView(generics.ListCreateAPIView):
    queryset = RechargePlan.objects.filter(is_deleted=False)
    serializer_class = RechargePlanSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]


#  Plan Detail
class RechargePlanDetailAPIView(generics.RetrieveUpdateAPIView):
    queryset = RechargePlan.objects.filter(is_deleted=False)
    serializer_class = RechargePlanSerializer
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]


#  Plan Soft Delete
class RechargePlanDeleteAPIView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def delete(self, request, pk):
        try:
            plan = RechargePlan.objects.get(pk=pk, is_deleted=False)
            plan.is_deleted = True
            plan.save()
            return Response({"message": "Plan deleted successfully"}, status=status.HTTP_200_OK)
        except RechargePlan.DoesNotExist:
            return Response({"error": "Plan not found"}, status=status.HTTP_404_NOT_FOUND)

from rest_framework import status, permissions
from payments.models import UserRecharge
from .services import RazorpayService
from .exceptions import PaymentException
import logging

logger = logging.getLogger('payments')

class RechargePlansView(APIView):
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        plans = RechargePlan.objects.filter(is_active=True, is_deleted=False)
        serializer = RechargePlanSerializer(plans, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


razorpay_client = razorpay.Client(
    auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET)
)

class UserRechargeView(APIView):
    """
    API endpoint to initiate a recharge for a user
    Creates Razorpay order and UserRecharge record
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        user = request.user
        plan_id = request.data.get("plan_id")

        if not plan_id:
            return Response(
                {"error": "plan_id is required"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            plan = RechargePlan.objects.get(id=plan_id, is_active=True, is_deleted=False)
        except RechargePlan.DoesNotExist:
            return Response(
                {"error": "Invalid or inactive recharge plan"},
                status=status.HTTP_404_NOT_FOUND
            )

        try:
            # Use service layer to create order
            order_data = RazorpayService.create_order(user, plan)
            
            return Response({
                "message": "Razorpay order created successfully",
                **order_data
            }, status=status.HTTP_200_OK)
            
        except PaymentException as e:
            logger.error(f"Payment error for user {user.id}: {str(e)}")
            return Response(
                {
                    "error": str(e),
                    "error_type": "PaymentException"
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as e:
            logger.error(f"Unexpected error in recharge initiation: {type(e).__name__} - {str(e)}", exc_info=True)
            
            # In DEBUG mode, provide detailed error for debugging
            error_response = {"error": "Failed to create payment order. Please try again."}
            if settings.DEBUG:
                error_response["error_details"] = str(e)
                error_response["error_type"] = type(e).__name__
                
            return Response(
                error_response,
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    

import hmac
import hashlib

class VerifyRechargePaymentView(APIView):
    """
    API endpoint to verify payment after user completes Razorpay checkout
    Verifies signature and updates recharge status
    """
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        user = request.user
        razorpay_order_id = request.data.get("razorpay_order_id")
        razorpay_payment_id = request.data.get("razorpay_payment_id")
        razorpay_signature = request.data.get("razorpay_signature")

        if not all([razorpay_order_id, razorpay_payment_id, razorpay_signature]):
            return Response(
                {"error": "Missing required payment details"},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Use service layer to process payment
            from .exceptions import (
                InvalidSignatureException,
                OrderNotFoundException,
                PaymentAlreadyProcessedException
            )
            
            result = RazorpayService.process_successful_payment(
                razorpay_order_id,
                razorpay_payment_id,
                razorpay_signature
            )
            
            return Response(result, status=status.HTTP_200_OK)
            
        except InvalidSignatureException as e:
            logger.warning(f"Invalid signature for order {razorpay_order_id}")
            return Response(
                {"error": "Payment verification failed. Invalid signature."},
                status=status.HTTP_400_BAD_REQUEST
            )
        except OrderNotFoundException as e:
            logger.error(f"Order not found: {razorpay_order_id}")
            return Response(
                {"error": "Recharge order not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except PaymentAlreadyProcessedException as e:
            # Return success for idempotency
            try:
                recharge = UserRecharge.objects.get(razorpay_order_id=razorpay_order_id)
                return Response({
                    "message": "Payment already processed",
                    "coins_added": recharge.coins_added,
                    "amount_paid": float(recharge.amount_paid),
                    "current_coin_balance": recharge.user.stats.coin_balance
                }, status=status.HTTP_200_OK)
            except UserRecharge.DoesNotExist:
                return Response(
                    {"error": "Payment already processed but details not found"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )
        except Exception as e:
            logger.error(f"Payment verification error: {str(e)}", exc_info=True)
            return Response(
                {"error": "Payment verification failed. Please contact support."},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class RazorpayWebhookView(APIView):
    """
    Webhook endpoint for Razorpay payment events
    
    Handles:
    - payment.captured: Payment successful
    - payment.failed: Payment failed
    - order.paid: Order completed
    """
    permission_classes = [permissions.AllowAny]
    authentication_classes = []

    def post(self, request):
        # Get webhook signature
        signature = request.headers.get('X-Razorpay-Signature')
        if not signature:
            logger.warning("Webhook received without signature")
            return Response(
                {"error": "Missing signature"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get raw request body for signature verification
        webhook_body = request.body.decode('utf-8')
        
        # Verify signature
        if not RazorpayService.verify_webhook_signature(webhook_body, signature):
            logger.warning("Webhook signature verification failed")
            return Response(
                {"error": "Invalid signature"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Extract event details
        payload = request.data
        event_type = payload.get('event')
        
        if not event_type:
            logger.error("Webhook received without event type")
            return Response(
                {"error": "Missing event type"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        logger.info(f"Processing webhook event: {event_type}")
        
        try:
            # Process webhook using service layer
            from .services import PaymentWebhookService
            PaymentWebhookService.handle_webhook(event_type, payload)
            
            return Response({"status": "ok"}, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Webhook processing failed: {str(e)}", exc_info=True)
            # Return 200 to prevent Razorpay from retrying
            # Error is logged for manual investigation
            return Response({"status": "error", "message": str(e)}, status=status.HTTP_200_OK)


class RedemptionOptionListCreateAPIView(APIView):
    permission_classes=[IsAdminUser]
    authentication_classes=[JWTAuthentication]

    def get(self, request):
        options = RedemptionOption.objects.filter(is_deleted=False)
        serializer = RedemptionOptionSerializer(options, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def post(self, request):
        serializer = RedemptionOptionSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class RedemptionOptionDetailAPIView(APIView):
    permission_classes = [IsAdminUser]
    authentication_classes = [JWTAuthentication]

    def get_object(self, pk):
        try:
            return RedemptionOption.objects.get(pk=pk, is_deleted=False)
        except RedemptionOption.DoesNotExist:
            return None

    def get(self, request, pk):
        option = self.get_object(pk)
        if not option:
            return Response({"detail": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = RedemptionOptionSerializer(option)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def put(self, request, pk):
        option = self.get_object(pk)
        if not option:
            return Response({"detail": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = RedemptionOptionSerializer(option, data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        option = self.get_object(pk)
        if not option:
            return Response({"detail": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = RedemptionOptionSerializer(option, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_200_OK)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        option = self.get_object(pk)
        if not option:
            return Response({"detail": "Not found"}, status=status.HTTP_404_NOT_FOUND)
        option.is_deleted = True
        option.save()
        return Response({"detail": "Deleted successfully"}, status=status.HTTP_204_NO_CONTENT)
    
class RedemptionOptionListViewExecutive(APIView):
    permission_classes=[IsAuthenticated]
    authentication_classes=[ExecutiveTokenAuthentication]

    def get(self, request):
        options = RedemptionOption.objects.filter(is_deleted=False)
        serializer = RedemptionOptionSerializer(options, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
class ExecutiveRedeemAPIView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [ExecutiveTokenAuthentication]  

    def post(self, request):
        executive = getattr(request, "user", None)  

        if not executive:
            return Response({"detail": "Executive not found or not authenticated."},
                            status=status.HTTP_401_UNAUTHORIZED)

        stats = getattr(executive, "stats", None)
        if not stats:
            return Response({"detail": "Executive stats not found."}, status=status.HTTP_404_NOT_FOUND)

        serializer = ExecutiveRedeemSerializer(data=request.data)
        if serializer.is_valid():
            redemption_option = serializer.validated_data["redemption_option"]
            redeem_amount = redemption_option.amount

            if stats.pending_payout < redeem_amount:
                return Response({"detail": "Insufficient pending payout to redeem this amount."},
                                status=status.HTTP_400_BAD_REQUEST)

            stats.pending_payout -= Decimal(redeem_amount)
            stats.save()

            redeem_request = ExecutivePayoutRedeem.objects.create(
                executive=executive,
                redemption_option=redemption_option,
                status="pending",
                upi_details=serializer.validated_data.get("upi_details"),
                account_number=serializer.validated_data.get("account_number"),
                ifsc_code=serializer.validated_data.get("ifsc_code")
            )

            return Response({
                "detail": "Redemption request created successfully.",
                "request_id": redeem_request.id,
                "amount": redeem_amount,
                "status": redeem_request.status,
                "executiveId": redeem_request.executive.id
            }, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
class ExecutiveRedeemHistoryAPIView(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [ExecutiveTokenAuthentication]

    def get(self, request):
        executive = getattr(request, "user", None)

        if not executive:
            return Response({"detail": "Executive not found or not authenticated."},
                            status=401)

        redeems = ExecutivePayoutRedeem.objects.filter(executive=executive).order_by("-requested_at")
        serializer = ExecutiveRedeemHistorySerializer(redeems, many=True)

        return Response(serializer.data)


class AdminRedeemListUpdateAPIView(APIView):
    permission_classes = [IsAdminUser]
    authentication_classes = [JWTAuthentication]

    # List all redemption requests
    def get(self, request):
        redeems = ExecutivePayoutRedeem.objects.all().order_by("-requested_at")
        serializer = AdminRedeemManageSerializer(redeems, many=True)
        return Response(serializer.data)

    # Update redemption request (approve/reject/paid)
    def patch(self, request, pk):
        try:
            redeem_request = ExecutivePayoutRedeem.objects.get(pk=pk)
        except ExecutivePayoutRedeem.DoesNotExist:
            return Response({"detail": "Redeem request not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = AdminRedeemManageSerializer(redeem_request, data=request.data, partial=True)
        if serializer.is_valid():
            redeem = serializer.save()

            if redeem.status in ["approved", "rejected", "paid"]:
                redeem.processed_at = now()
                redeem.save()

            return Response(AdminRedeemManageSerializer(redeem).data)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
class UserRechargeHistoryViewAdmin(APIView):
    permission_classes = [IsAdminUser]
    authentication_classes=[JWTAuthentication]  

    def get(self, request, user_id):
        try:
            user = UserProfile.objects.get(id=user_id)
        except UserProfile.DoesNotExist:
            return Response({"message": "User not found"}, status=status.HTTP_404_NOT_FOUND)

        recharges = UserRecharge.objects.filter(user=user).order_by('-created_at')
        serializer = UserRechargeSerializer(recharges, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
class UserRechargeHistoryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        try:
            user = request.user

            recharges = UserRecharge.objects.filter(user=user).order_by('-created_at')
            serializer = UserRechargeSerializer(recharges, many=True)

            return Response(serializer.data, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error fetching recharge history: {type(e).__name__} - {str(e)}", exc_info=True)
            return Response(
                {"error": "Failed to fetch recharge history"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
from rest_framework import status
from django.db import transaction

class AdminRechargeView(APIView):
    permission_classes = []
    authentication_classes = []

    def post(self, request):
        user_id = request.data.get("user_id")
        plan_id = request.data.get("plan_id")
        coins_to_add = request.data.get("coins_added")
        amount_paid = request.data.get("amount_paid")

        if not user_id:
            return Response({"error": "user_id is required"}, status=status.HTTP_400_BAD_REQUEST)

        # Fetch user instance
        try:
            user = UserProfile.objects.get(id=user_id)
        except UserProfile.DoesNotExist:
            return Response({"error": "User not found"}, status=status.HTTP_404_NOT_FOUND)

        plan = None
        if plan_id:
            try:
                plan = RechargePlan.objects.get(id=plan_id)
                coins_to_add = plan.get_adjusted_coin_package() if coins_to_add is None else coins_to_add
                amount_paid = plan.calculate_final_price() if amount_paid is None else amount_paid
            except RechargePlan.DoesNotExist:
                return Response({"error": "Recharge plan not found"}, status=status.HTTP_404_NOT_FOUND)

        # Validate coins and amount
        if coins_to_add is None or amount_paid is None:
            return Response({"error": "coins_added and amount_paid are required if no plan_id is provided."}, status=status.HTTP_400_BAD_REQUEST)

        # Convert to proper types
        try:
            coins_to_add = int(coins_to_add)
            amount_paid = float(amount_paid)
        except ValueError:
            return Response({"error": "Invalid coins_added or amount_paid value."}, status=status.HTTP_400_BAD_REQUEST)

        # Use a transaction to safely update stats using select_for_update
        try:
            with transaction.atomic():
                # Lock the stats row
                stats = None
                if hasattr(user, "stats"):
                    stats = type(user.stats).objects.select_for_update().get(pk=user.stats.pk)

                # Create the recharge
                recharge = UserRecharge.objects.create(
                    user=user,
                    plan=plan,
                    coins_added=coins_to_add,
                    amount_paid=amount_paid,
                    is_successful=True,
                    by_admin=True,
                    payment_status="successful",
                )

                # Update user coin balance specifically
                if stats:
                    import logging
                    logger = logging.getLogger("payments")
                    old_balance = stats.coin_balance
                    stats.coin_balance += coins_to_add
                    stats.save(update_fields=["coin_balance"])
                    logger.info(f"Admin Recharge: User {user.id} balance changed from {old_balance} to {stats.coin_balance} by adding {coins_to_add}")

            return Response({
                "message": f"Recharge successful for user {user.name or user.user_id}",
                "coins_added": coins_to_add,
                "amount_paid": amount_paid,
                "current_coin_balance": stats.coin_balance if stats else 0,
            }, status=status.HTTP_200_OK)
        except Exception as e:
            import logging
            logger = logging.getLogger("payments")
            logger.error(f"Error during admin recharge for user {user.id}: {str(e)}")
            return Response({"error": "Failed to process recharge securely."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
from django.utils import timezone
from django.db.models import Sum


class RechargeAnalyticsView(APIView):
    permission_classes = [IsAdminUser]  
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        try:
            today = timezone.now().date()

            successful_user_recharges = UserRecharge.objects.filter(
                is_successful=True,
                by_admin=False
            )

            today_stats = successful_user_recharges.filter(
                created_at__date=today
            ).aggregate(
                today_coins=Sum('coins_added'),
                today_revenue=Sum('amount_paid')
            )
            today_coin_sales = today_stats['today_coins'] or 0
            today_revenue = float(today_stats['today_revenue'] or 0)

            total_coin_sales = successful_user_recharges.aggregate(
                total_coins=Sum('coins_added')
            )['total_coins'] or 0

            total_revenue = float(successful_user_recharges.aggregate(
                total_amount=Sum('amount_paid')
            )['total_amount'] or 0)

            admin_recharges = UserRecharge.objects.filter(
                is_successful=True,
                by_admin=True
            )

            admin_spent_amount = float(admin_recharges.aggregate(
                total_amount=Sum('amount_paid')
            )['total_amount'] or 0)

            admin_coins_spent = admin_recharges.aggregate(
                total_coins=Sum('coins_added')
            )['total_coins'] or 0

            data = {
                "today_coin_sales": today_coin_sales,
                "today_revenue": today_revenue,
                "total_coin_sales": total_coin_sales,
                "total_revenue": total_revenue,
                "admin_spent_amount": admin_spent_amount,
                "admin_coins_spent": admin_coins_spent
            }

            return Response(data, status=200)

        except Exception as e:
            return Response({"error": str(e)}, status=500)
        
class UserRechargeListView(APIView):
    permission_classes = [IsAdminUser]
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        user_id = request.query_params.get("user_id")
        status_filter = request.query_params.get("status")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")

        recharges = UserRecharge.objects.select_related("user", "plan").order_by("-created_at")

        if user_id:
            recharges = recharges.filter(user__id=user_id)
        if status_filter:
            recharges = recharges.filter(payment_status=status_filter)
        if start_date and end_date:
            recharges = recharges.filter(created_at__range=[start_date, end_date])

        paginator = CustomUserPagination()
        paginated_qs = paginator.paginate_queryset(recharges, request)
        serializer = UserRechargeSerializer(paginated_qs, many=True)

        return paginator.get_paginated_response(serializer.data)


class UserRechargeExportExcelView(APIView):
    """
    Exports the full recharge list (respecting the same filters as
    UserRechargeListView) as a formatted Excel (.xlsx) file.
    """
    permission_classes = [IsAdminUser]
    authentication_classes = [JWTAuthentication]

    def get(self, request):
        user_id = request.query_params.get("user_id")
        status_filter = request.query_params.get("status")
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")

        recharges = UserRecharge.objects.select_related("user", "plan").order_by("-created_at")

        if user_id:
            recharges = recharges.filter(user__id=user_id)
        if status_filter:
            recharges = recharges.filter(payment_status=status_filter)
        if start_date and end_date:
            recharges = recharges.filter(created_at__range=[start_date, end_date])

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Recharges"

        headers = [
            "ID",
            "User ID",
            "User Name",
            "Plan Name",
            "Coins Added",
            "Amount Paid (₹)",
            "Payment Status",
            "Successful",
            "By Admin",
            "Razorpay Order ID",
            "Razorpay Payment ID",
            "Created At",
            "Updated At",
        ]

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        sheet.append(headers)
        for col_num, _ in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        sheet.freeze_panes = "A2"
        sheet.row_dimensions[1].height = 28

        status_fills = {
            "successful": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
            "pending": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "failed": PatternFill(start_color="FCE4E4", end_color="FCE4E4", fill_type="solid"),
        }

        total_amount = Decimal("0")
        total_coins = 0
        row_num = 2

        for recharge in recharges.iterator():
            row = [
                recharge.id,
                recharge.user.user_id if recharge.user else "",
                recharge.user.name if recharge.user else "",
                recharge.plan.plan_name if recharge.plan else "",
                recharge.coins_added,
                float(recharge.amount_paid),
                recharge.payment_status.title(),
                "Yes" if recharge.is_successful else "No",
                "Yes" if recharge.by_admin else "No",
                recharge.razorpay_order_id or "",
                recharge.razorpay_payment_id or "",
                recharge.created_at.strftime("%Y-%m-%d %H:%M:%S") if recharge.created_at else "",
                recharge.updated_at.strftime("%Y-%m-%d %H:%M:%S") if recharge.updated_at else "",
            ]
            sheet.append(row)

            fill = status_fills.get(recharge.payment_status)
            for col_num in range(1, len(headers) + 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.border = thin_border
                if fill:
                    cell.fill = fill
                if col_num == 6:
                    cell.number_format = "#,##0.00"

            if recharge.is_successful:
                total_amount += recharge.amount_paid
                total_coins += recharge.coins_added

            row_num += 1

        # Totals row (successful recharges only)
        totals_font = Font(bold=True)
        totals_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        sheet.append([
            "", "", "", "Total (Successful)", total_coins, float(total_amount),
            "", "", "", "", "", "", ""
        ])
        for col_num in range(1, len(headers) + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.font = totals_font
            cell.fill = totals_fill
            cell.border = thin_border
            if col_num == 6:
                cell.number_format = "#,##0.00"

        column_widths = [8, 14, 22, 22, 14, 16, 16, 12, 12, 26, 26, 20, 20]
        for i, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[get_column_letter(i)].width = width

        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_num - 1}"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"recharges_{now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response